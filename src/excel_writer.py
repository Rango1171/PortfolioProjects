"""
Writes job search results to a dated Excel file using openpyxl.
Each row represents one job. Hyperlinks are embedded for URLs.
"""

import logging
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADERS = [
    "#",
    "Date Found",
    "Company",
    "Role Title",
    "Country",
    "Job Description Link",
    "LinkedIn Link",
    "Visa Sponsorship",
    "Match Score",
    "Tailored Resume (Drive)",
    "Notes",
]

COLUMN_WIDTHS = [5, 14, 28, 32, 12, 45, 45, 18, 14, 45, 30]

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LINK_FONT = Font(name="Calibri", color="0563C1", underline="single", size=10)
BODY_FONT = Font(name="Calibri", size=10)

SPONSORSHIP_COLORS = {
    "Yes": "C6EFCE",   # green
    "No": "FFC7CE",    # red
    "Unknown": "FFEB9C",  # yellow
}

def _score_color(score: int) -> str:
    if score >= 80:
        return "C6EFCE"   # green
    if score >= 60:
        return "FFEB9C"   # yellow
    return "FFC7CE"        # red

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _get_or_create_workbook(path: Path) -> tuple[Workbook, any]:
    if path.exists():
        wb = load_workbook(str(path))
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = f"Jobs {date.today().isoformat()}"

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    for col_idx, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb, ws


def _next_row(ws) -> int:
    return ws.max_row + 1 if ws.max_row > 1 else 2


def _hyperlink_cell(ws, row: int, col: int, url: str, label: str = "") -> None:
    label = label or url or "N/A"
    cell = ws.cell(row=row, column=col, value=label if url else "N/A")
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    else:
        cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cell.border = THIN_BORDER


def append_job(ws, row_num: int, job_row: dict) -> None:
    is_alt = row_num % 2 == 0
    bg_fill = ALT_FILL if is_alt else None

    def set_cell(col, value, is_link=False, url=""):
        cell = ws.cell(row=row_num, column=col, value=value)
        if is_link and url:
            cell.hyperlink = url
            cell.font = LINK_FONT
        else:
            cell.font = BODY_FONT
        if bg_fill and not is_link:
            cell.fill = bg_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
        return cell

    set_cell(1, job_row["index"])
    set_cell(2, job_row["date_found"])
    set_cell(3, job_row["company"])
    set_cell(4, job_row["title"])
    set_cell(5, job_row["country"])

    # Job description link
    jd_url = job_row.get("job_url", "")
    _hyperlink_cell(ws, row_num, 6, jd_url, "View Job")

    # LinkedIn link
    li_url = job_row.get("linkedin_url", "")
    _hyperlink_cell(ws, row_num, 7, li_url, "LinkedIn" if li_url else "N/A")

    # Visa sponsorship with color coding
    sponsorship = job_row.get("visa_sponsorship", "Unknown")
    cell = ws.cell(row=row_num, column=8, value=sponsorship)
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    color = SPONSORSHIP_COLORS.get(sponsorship, "FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Match score with color coding
    score = job_row.get("resume_match_score", 0)
    score_cell = ws.cell(row=row_num, column=9, value=f"{score}%")
    score_cell.font = BODY_FONT
    score_cell.alignment = Alignment(horizontal="center", vertical="center")
    score_cell.border = THIN_BORDER
    sc = _score_color(score)
    score_cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")

    # Tailored resume Drive link
    drive_url = job_row.get("drive_link", "")
    _hyperlink_cell(ws, row_num, 10, drive_url, "Open Resume" if drive_url else "N/A")

    # Notes (empty, user fills)
    set_cell(11, "")
    ws.row_dimensions[row_num].height = 16


class ExcelWriter:
    def __init__(self, excel_dir: str):
        self.excel_dir = Path(excel_dir)
        self.excel_dir.mkdir(parents=True, exist_ok=True)
        today = date.today().isoformat()
        self.path = self.excel_dir / f"jobs_{today}.xlsx"
        self.wb, self.ws = _get_or_create_workbook(self.path)
        self._row = _next_row(self.ws)

    def add_job(self, job_row: dict) -> None:
        job_row["index"] = self._row - 1  # 1-based job counter
        append_job(self.ws, self._row, job_row)
        self._row += 1

    def save(self) -> Path:
        self.wb.save(str(self.path))
        logger.info(f"Excel saved: {self.path}")
        return self.path
